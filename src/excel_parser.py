"""Excel parser for video clip definitions."""
import logging
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

# Setup logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

# Console handler for debugging
_console_handler = logging.StreamHandler()
_console_handler.setLevel(logging.DEBUG)
_console_handler.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
logger.addHandler(_console_handler)


@dataclass
class ClipDefinition:
    """Definition of a video clip to extract."""
    start_time: datetime
    end_time: datetime
    description: str
    
    @property
    def duration_seconds(self) -> float:
        """Calculate clip duration in seconds."""
        return (self.end_time - self.start_time).total_seconds()
    
    def get_output_filename(self) -> str:
        """Generate output filename from description."""
        # Clean description for filename
        clean_desc = self.description.strip()
        # Remove invalid filename characters
        for char in '<>:"/\\|?*':
            clean_desc = clean_desc.replace(char, '_')
        return f"{clean_desc}.mp4"


def parse_time_string(time_str: str) -> datetime | None:
    """
    Parse time string in various formats.
    
    Supports:
    - "起 2026-01-15 10:45:02 止 2026-01-15 11:30:00"
    - "2026-01-15 10:45:02"
    """
    import re
    
    # Pattern for datetime
    pattern = r'(\d{4})[-.](\d{2})[-.](\d{2})[\s_](\d{2}):(\d{2}):(\d{2})'
    match = re.search(pattern, time_str)
    
    if match:
        try:
            return datetime(
                int(match.group(1)), int(match.group(2)), int(match.group(3)),
                int(match.group(4)), int(match.group(5)), int(match.group(6))
            )
        except ValueError:
            return None
    
    return None


def parse_excel_clips(excel_path: Path, debug: bool = False) -> list[ClipDefinition]:
    """
    Parse Excel file to extract clip definitions.
    
    Expected columns:
    - "起始时间点" (Start time)
    - "问题描述" (Description)
    
    The start time column contains strings like:
    "起 yyyy-mm-dd hh:mm:ss 止 yyyy-mm-dd hh:mm:ss"
    
    Args:
        excel_path: Path to Excel file
        debug: Enable detailed debug logging
    
    Returns:
        List of ClipDefinition objects
    """
    clips = []
    
    if debug:
        logger.setLevel(logging.DEBUG)
        logger.info(f"开始解析 Excel 文件: {excel_path}")
    
    try:
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        
        logger.info(f"工作表名称: {sheet.title}")
        logger.info(f"总行数: {sheet.max_row}, 总列数: {sheet.max_column}")
        
        # Find column indices
        header_row = 1
        time_col = None
        desc_col = None
        
        # Log all headers
        if debug:
            logger.debug("=== 表头内容 ===")
            for col_idx, cell in enumerate(sheet[header_row], start=1):
                logger.debug(f"  列 {col_idx}: '{cell.value}'")
        
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            cell_value = str(cell.value or "").strip()
            # Match time columns: 起止时间, 起始时间, 开始时间, 时间
            if "时间" in cell_value and ("起" in cell_value or "止" in cell_value or "开始" in cell_value):
                time_col = col_idx
                logger.debug(f"找到时间列 (模糊匹配): 列 {col_idx} = '{cell_value}'")
            elif "时间" in cell_value and "问题" not in cell_value:
                # Also match just "时间" if not combined with "问题"
                time_col = col_idx
                logger.debug(f"找到时间列 (模糊匹配): 列 {col_idx} = '{cell_value}'")
            elif "问题" in cell_value or "描述" in cell_value or "标题" in cell_value:
                desc_col = col_idx
                logger.debug(f"找到描述列 (模糊匹配): 列 {col_idx} = '{cell_value}'")
        
        if not time_col or not desc_col:
            logger.warning("模糊匹配未找到列，尝试精确匹配...")
            # Try common column names in Chinese
            for col_idx, cell in enumerate(sheet[header_row], start=1):
                cell_value = str(cell.value or "").strip().lower()
                if cell_value in ["起止时间点", "起止时间", "起始时间点", "起始时间", "开始时间", "时间"]:
                    time_col = col_idx
                    logger.debug(f"找到时间列 (精确匹配): 列 {col_idx} = '{cell_value}'")
                elif cell_value in ["问题描述", "描述", "标题", "片段名称"]:
                    desc_col = col_idx
                    logger.debug(f"找到描述列 (精确匹配): 列 {col_idx} = '{cell_value}'")
        
        if not time_col or not desc_col:
            logger.error(f"未找到必需列: time_col={time_col}, desc_col={desc_col}")
            logger.error("请确保表头包含 '起始时间' 和 '问题描述' 相关字段")
            workbook.close()
            return clips
        
        logger.info(f"列映射: 时间列={time_col}, 描述列={desc_col}")
        
        # Parse data rows
        import re
        
        # Flexible datetime pattern - supports multiple date/time formats
        # 2026-01-15 10:45:02, 2026/01/15 10:45:02, 2026.01.15 10:45:02, etc.
        # Also handles: 2026-1-5 10:45:02 (single digit month/day)
        
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            time_cell = sheet.cell(row=row_idx, column=time_col)
            desc_cell = sheet.cell(row=row_idx, column=desc_col)
            
            time_str = str(time_cell.value or "").strip()
            desc = str(desc_cell.value or "").strip()
            
            if debug:
                logger.debug(f"=== 第 {row_idx} 行 ===")
                logger.debug(f"  时间单元格原始值: {repr(time_cell.value)}")
                logger.debug(f"  时间字符串(trim后): '{time_str}'")
                logger.debug(f"  描述: '{desc}'")
            
            if not time_str or not desc:
                if debug:
                    logger.debug(f"  跳过: 时间或描述为空")
                continue
            
            # Normalize time string: handle various separators
            # Handle cases like: "起 xxx 止 xxx", "起 xxx\n止 xxx", "起 xxx\r\n止 xxx", "起 xxx止 xxx"
            normalized = time_str
            
            # Replace common separators before "止" with a single space
            # Handles: \n止, \r\n止, \r止, "止" directly, " 止"
            import re as re_module
            normalized = re_module.sub(r'[\r\n]+', ' ', normalized)  # Replace all line breaks with space
            normalized = re_module.sub(r'\s+', ' ', normalized)      # Normalize multiple spaces
            normalized = re_module.sub(r'\s*止\s*', ' 止 ', normalized)  # Ensure "止" has spaces around
            
            if debug and normalized != time_str:
                logger.debug(f"  规范化后的时间字符串: '{normalized}'")
            
            # Extract all datetime patterns - supports multiple formats
            # 2026-01-15 10:45:02, 2026/01/15 10:45:02, 2026.01.15 10:45:02
            # Also handles single-digit: 2026-1-5 10:45:02
            pattern = r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})[\s_]+(\d{1,2}):(\d{1,2}):(\d{1,2})'
            matches = re.findall(pattern, normalized)
            
            if debug:
                logger.debug(f"  正则匹配结果: 找到 {len(matches)} 个时间戳")
                for i, m in enumerate(matches):
                    logger.debug(f"    时间戳 {i+1}: {m[0]}-{m[1]}-{m[2]} {m[3]}:{m[4]}:{m[5]}")
            
            if len(matches) >= 2:
                try:
                    start_time = datetime(
                        int(matches[0][0]), int(matches[0][1]), int(matches[0][2]),
                        int(matches[0][3]), int(matches[0][4]), int(matches[0][5])
                    )
                    end_time = datetime(
                        int(matches[1][0]), int(matches[1][1]), int(matches[1][2]),
                        int(matches[1][3]), int(matches[1][4]), int(matches[1][5])
                    )
                    
                    clips.append(ClipDefinition(
                        start_time=start_time,
                        end_time=end_time,
                        description=desc
                    ))
                    logger.info(f"成功解析片段: {desc} ({start_time} ~ {end_time})")
                except ValueError as ve:
                    logger.warning(f"第 {row_idx} 行时间解析失败: {ve}")
                    continue
            else:
                logger.warning(f"第 {row_idx} 行: 时间格式不匹配，需要 2 个时间戳，找到 {len(matches)} 个")
                logger.warning(f"  请确保格式类似: '起 2026-01-15 10:45:02 止 2026-01-15 11:30:00'")
                if debug and len(matches) == 1:
                    logger.warning(f"  只找到 1 个时间戳，可能缺少结束时间")
                elif debug and len(matches) == 0:
                    logger.warning(f"  未找到任何时间戳，请检查日期时间格式是否正确")
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"解析 Excel 时出错: {e}")
        import traceback
        logger.error(traceback.format_exc())
    
    logger.info(f"解析完成: 共提取 {len(clips)} 个片段")
    return clips
